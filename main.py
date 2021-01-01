import os
import json
import datetime

from openpyxl import load_workbook
from util import download_file


def main():
    #宮城県のデータのexcelファイル（公式）
    url="https://www.pref.miyagi.jp/uploaded/attachment/826977.xlsx"
    file_name=download_file(url)

    wb=load_workbook(file_name,data_only=True)
    ws=wb["日別集計（HP掲載）"]

    #データを集めていく
    today=datetime.date.today()
    d=[]
    max_row=ws.max_row
    for i in range(2,max_row):
        dt=ws.cell(i,1).value #日付

        #cellに記入されていない行に来たら終了
        if dt is None:
            break
        #今日以降のセルの行に行くと終了
        date=dt.date()
        if date>today:
            break

        date_isoformat=date.isoformat()
        count=ws.cell(i,7).value #合計
        #今日の日付の分が0人の時は、記入されていないこととして、ループ終了する
        if date==today and count==0:
            break
            
        print(date,count)
        d.append({
            "date":date_isoformat,
            "count":count
        })
    
    #jsonに書きだし
    data={
        "data":d
    }
    print(data)
    os.makedirs("data",exist_ok=True)
    with open("data/miyagi_data.json","w") as f:
        json.dump(data,f,indent=4)

if __name__ == "__main__":
    main()
